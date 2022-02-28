import os
from os import path
import pandas as pd
from openpyxl.workbook import Workbook
from openpyxl.styles import Alignment, Border, Side

# Create folder is non exists
if not path.exists('Output'):
    os.mkdir('Output')

# Get file
file = 'Waste Sort.xls'
file_sheets = pd.ExcelFile(file).sheet_names

# Read in file 
data = pd.read_excel(file, header=0)
data_list = data.to_numpy()
print(data_list)

# Strip all white space
strip_data = []

for x in data_list:
    a = x[0].strip()
    b = x[1]
    c = x[2].strip()
    d = x[3]
    e = x[4]
    f = x[5]
    striped = [a, b, c, d, e, f]
    strip_data.append(striped)
print(strip_data)

# Separate between stocktake and expired
stock_take_list = []
date_expired_list = []

for item in strip_data:
    if item[2] == "DATE EXPIRED":
        date_expired_list.append(item)
    else:
        stock_take_list.append(item)

print(stock_take_list)
print(date_expired_list)

# Create workbook for stock take
wb = Workbook()
ws = wb.active
ws.title = 'Stocktake Waste'
print('Create Excel Workbook - Stocktake Waste')

# Set Column Width 
ws.column_dimensions['A'].width = 32
ws.column_dimensions['B'].width = 11
ws.column_dimensions['C'].width = 15
ws.column_dimensions['D'].width = 9
ws.column_dimensions['E'].width = 13
ws.column_dimensions['F'].width = 13

# Set row names
print('Create Column Names')
ws['A1'] = 'Stock Description'
ws['B1'] = 'Date'
ws['C1'] = 'Detail'
ws['D1'] = 'Qty Move'
ws['E1'] = 'Val at Avg Cost'
ws['F1'] = 'Val at Lat Cost'

 # # loop through sheet totals
i = 2
while i <= len(stock_take_list) :
    for item in stock_take_list:
        ws[f'A{i}'] = item[0]
        ws[f'B{i}'] = item[1]
        ws[f'C{i}'] = item[2]
        ws[f'D{i}'] = item[3]
        ws[f'E{i}'] = item[4]
        ws[f'F{i}'] = item[5]
        print(item[0])
        print(item[1])
        print(item[2])
        print(item[3])
        print(item[4])
        print(item[5])
        
        ws[f'D{i}'].number_format = '# ##0.00'
        ws[f'E{i}'].number_format = '# ##0.00'
        ws[f'F{i}'].number_format = '# ##0.00'
        ws[f'B{i}'].number_format = 'DD/MM/YYYY'
        i += 1

wb.save('Output/Stocktake Waste.xlsx')

# Create workbook for date expired
wb = Workbook()
ws = wb.active
ws.title = 'Waste'
print('Create Excel Workbook - Waste')

# Set Column Width 
ws.column_dimensions['A'].width = 32
ws.column_dimensions['B'].width = 11
ws.column_dimensions['C'].width = 15
ws.column_dimensions['D'].width = 9
ws.column_dimensions['E'].width = 13
ws.column_dimensions['F'].width = 13

# Set row names
print('Create Column Names')
ws['A1'] = 'Stock Description'
ws['B1'] = 'Date'
ws['C1'] = 'Detail'
ws['D1'] = 'Qty Move'
ws['E1'] = 'Val at Avg Cost'
ws['F1'] = 'Val at Lat Cost'

 # # loop through sheet totals
i = 2
while i <= len(date_expired_list) :
    for item in date_expired_list:
        ws[f'A{i}'] = item[0]
        ws[f'B{i}'] = item[1]
        ws[f'C{i}'] = item[2]
        ws[f'D{i}'] = item[3]
        ws[f'E{i}'] = item[4]
        ws[f'F{i}'] = item[5]
        print(item[0])
        print(item[1])
        print(item[2])
        print(item[3])
        print(item[4])
        print(item[5])
        
        ws[f'D{i}'].number_format = '# ##0.00'
        ws[f'E{i}'].number_format = '# ##0.00'
        ws[f'F{i}'].number_format = '# ##0.00'
        ws[f'B{i}'].number_format = 'DD/MM/YYYY'
        i += 1

wb.save('Output/Waste.xlsx')

print('...')
print('...')
print('Save Workbook - Stocktake Waste')
print('Save Workbook - Waste')



